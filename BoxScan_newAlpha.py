import sys
import os
from pathlib import Path
import re
import json
import csv
from datetime import datetime

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QLabel, QLineEdit,
    QPushButton, QVBoxLayout, QHBoxLayout, QGridLayout, QGroupBox,
    QMessageBox, QFileDialog, QInputDialog, QTextEdit,
    QTreeWidget, QTreeWidgetItem, QMenu, QAction, QHeaderView,
    QToolTip, QCheckBox, QScrollArea, QScrollBar, QMenuBar, QActionGroup,
    QStyleFactory, QDialog, QSpacerItem, QSizePolicy
)
from PyQt5.QtGui import QIcon, QFont, QClipboard, QPixmap, QColor
from PyQt5.QtCore import Qt, pyqtSignal, QObject, QTimer, QEvent

import openpyxl
from openpyxl.styles import Alignment

import pyzbar.pyzbar as pyzbar
import threading
import pyperclip


class ToolTip(QObject):
    def __init__(self, widget):
        super().__init__()
        self.widget = widget
        self.tipwindow = None
        self.id = None
        self.x = 0
        self.y = 0
        self.delay = 500
        self.widget.installEventFilter(self)

    def showtip(self, text):
        self.text = text
        if self.tipwindow or not self.text:
            return

        def show_delayed_tip():
            if not self.tipwindow:
                rect = self.widget.rect()
                point = self.widget.mapToGlobal(rect.bottomLeft())
                x = point.x() + 5
                y = point.y() + 5

                self.tipwindow = tw = QDialog(self.widget, Qt.FramelessWindowHint | Qt.ToolTip)
                tw.setStyleSheet("QDialog {background-color: #ffffe0; border: 1px solid black;}")
                layout = QVBoxLayout(tw)
                layout.setContentsMargins(5, 5, 5, 5)
                label = QLabel(self.text, tw)
                label.setFont(QFont("Tahoma", 8))
                layout.addWidget(label)
                tw.adjustSize()
                tw.move(x, y)
                tw.show()

        self.id = self.widget.timerEvent(self.delay)
        QTimer.singleShot(self.delay, show_delayed_tip)

    def hidetip(self):
        if self.id:
            self.widget.killTimer(self.id)
            self.id = None
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.close()
            tw.destroy()

    def eventFilter(self, watched, event):
        if watched == self.widget:
            if event.type() == QEvent.Enter:
                ToolTip.showtip(self, self._tooltip_text)
            elif event.type() == QEvent.Leave:
                ToolTip.hidetip(self)
        return super().eventFilter(watched, event)

    def setToolTip(self, text):
        self._tooltip_text = text


class QBarcodeApp(QMainWindow):
    def __init__(self):
        super().__init__()
        print("__init__ started")  # DEBUG
        self.setWindowTitle("ScanBox")

        if getattr(sys, '_MEIPASS', None):
            base_path = sys._MEIPASS
        else:
            base_path = Path(__file__).resolve().parent

        resources_path = Path(base_path) / "resources"
        os.makedirs(resources_path, exist_ok=True)
        icon_path = str(resources_path / "icon.ico")

        if os.path.exists(icon_path):
            try:
                self.setWindowIcon(QIcon(icon_path))
            except Exception as e:
                print(f"Не удалось установить иконку: {e}")

        self.log_dir = os.path.join(base_path, "logs")
        os.makedirs(self.log_dir, exist_ok=True)

        self.all_boxes = {}
        self.current_box_barcode = ""
        self.search_query = ""
        # --- JSON State File Location: Hidden directory in user's home ---
        self.state_file_dir = Path(os.path.expanduser("~")) / ".ScanBox"
        os.makedirs(self.state_file_dir, exist_ok=True)
        self.state_file = str(self.state_file_dir / "barcode_app_state.json")
        print(f"State file path: {self.state_file}") # DEBUG


        self.box_bg_color = "#f2f2f2"
        self.history_file = None
        self.history_window = None
        self.history_tree = None
        self.comments = {}
        self.history_filter_query = ""

        self.COLOR_BG = "#f8f9fa"
        self.COLOR_FRAME_BG = "#ffffff"
        self.COLOR_ENTRY_BG = "#ffffff"
        self.COLOR_BUTTON_BG = "#e0f7fa"
        self.COLOR_BUTTON_FG = "#495057"
        self.COLOR_BUTTON_ACTIVE_BG = "#b2ebf2"
        self.COLOR_BUTTON_PRESSED_BG = "#80deea"
        self.COLOR_HEADER_BG = "#f8f9fa"
        self.COLOR_HEADER_FG = "#212529"
        self.COLOR_SCROLLBAR_BG = "#f1f3f5"
        self.COLOR_SCROLLBAR_TROUGH = "#ced4da"
        self.COLOR_SCROLLBAR_ARROW = "black"
        self.COLOR_SCROLLBAR_THUMB = "#949ca1"

        self.font_label = QFont("Segoe UI", 9)
        self.font_entry = QFont("Segoe UI", 9)
        self.font_button = QFont("Segoe UI Semibold", 10)
        self.font_treeview = QFont("Segoe UI", 9)
        self.font_treeview_heading = QFont("Segoe UI Semibold", 11)
        self.font_menu = QFont("Segoe UI", 10)

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.main_layout = QVBoxLayout()
        self.central_widget.setLayout(self.main_layout)
        self.main_layout.setContentsMargins(10, 10, 10, 10)

        self.strict_validation_enabled = True

        self.create_menu_bar()
        self.create_search_frame()
        self.create_box_frame()
        self.create_item_scan_frame()
        self.create_items_frame()
        self.create_control_frame()
        self.create_status_bar()

        self.load_state()

        self.clipboard = QApplication.clipboard()

        self.setStyleSheet(self.get_stylesheet())

        self.setGeometry(100, 100, 1280, 720)
        print("__init__ finished") # DEBUG

    def get_stylesheet(self):
        return f"""
            QMainWindow {{
                background-color: {self.COLOR_BG};
            }}
            QFrame {{
                background-color: {self.COLOR_FRAME_BG};
            }}
            QGroupBox {{
                background-color: {self.COLOR_FRAME_BG};
                border: 1px groove #ced4da;
                border-radius: 2px;
                margin-top: 0.5em;
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 3px 0 3px;
                font: bold 11pt "Segoe UI Semibold";
                color: {self.COLOR_HEADER_FG};
                background-color: {self.COLOR_FRAME_BG};
            }}
            QLabel {{
                font: 9pt "Segoe UI";
                background-color: {self.COLOR_BG};
            }}
            QLineEdit {{
                font: 9pt "Segoe UI";
                background-color: {self.COLOR_ENTRY_BG};
                border: 1px solid #ced4da;
                border-radius: 2px;
                padding: 2px;
            }}
            QPushButton {{
                font: bold 10pt "Segoe UI Semibold";
                background-color: {self.COLOR_BUTTON_BG};
                color: {self.COLOR_BUTTON_FG};
                border: none;
                padding: 5px 15px;
                border-radius: 3px;
            }}
            QPushButton:hover {{
                background-color: {self.COLOR_BUTTON_ACTIVE_BG};
            }}
            QPushButton:pressed {{
                background-color: {self.COLOR_BUTTON_PRESSED_BG};
            }}
            QMenuBar {{
                background-color: {self.COLOR_HEADER_BG};
                color: {self.COLOR_HEADER_FG};
                border-bottom: 1px solid #ced4da;
            }}
            QMenuBar::item {{
                background-color: transparent;
            }}
            QMenuBar::item:selected {{
                background-color: #e0e0e0;
            }}
            QMenu {{
                font: 10pt "Segoe UI";
                background-color: {self.COLOR_FRAME_BG};
                border: 1px solid #ced4da;
            }}
            QMenu::item:selected {{
                background-color: #bbdefb;
            }}
            QTreeWidget {{
                font: 9pt "Segoe UI";
                background-color: white;
                alternate-background-color: #f0f0f0;
                border: 1px solid #ced4da;
            }}
            QHeaderView::section {{
                background-color: {self.COLOR_HEADER_BG};
                font: bold 11pt "Segoe UI Semibold";
                border: none;
                border-bottom: 1px solid #ced4da;
                padding: 4px;
                qproperty-alignment: AlignCenter;
            }}
            QTreeWidget::item:selected {{
                background-color: #bbdefb;
                color: black;
            }}
            QScrollBar:vertical {{
                background-color: {self.COLOR_SCROLLBAR_BG};
                width: 10px;
                margin: 0px 0px 0px 0px;
            }}
            QScrollBar::handle:vertical {{
                background-color: {self.COLOR_SCROLLBAR_THUMB};
                min-height: 20px;
                border-radius: 5px;
            }}
            QScrollBar::add-line:vertical {{
                height: 0px;      subcontrol-position: bottom;
                subcontrol-origin: margin;
            }}
            QScrollBar::sub-line:vertical {{
                height: 0 px;
                subcontrol-position: top left;
                subcontrol-origin: margin;
            }}
            QScrollBar::up-arrow:vertical, QScrollBar::down-arrow:vertical {{
                background: none;
            }}
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {{
                background: none;
            }}
            QScrollBar:horizontal {{
                background-color: {self.COLOR_SCROLLBAR_BG};
                height: 10px;
                margin: 0px 0px 0px 0px;
            }}
            QScrollBar::handle:horizontal {{
                background-color: {self.COLOR_SCROLLBAR_THUMB};
                min-width: 20px;
                border-radius: 5px;
            }}
            QScrollBar::add-line:horizontal {{
                width: 0px;       subcontrol-position: right;
                subcontrol-origin: margin;
            }}
            QScrollBar::sub-line:horizontal {{
                width: 0 px;
                subcontrol-position: left top;
                subcontrol-origin: margin;
            }}
            QScrollBar::left-arrow:horizontal, QScrollBar::right-arrow:horizontal {{
                background: none;
            }}
            QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {{
                background: none;
            }}
            QStatusBar {{
                background-color: {self.COLOR_HEADER_BG};
                color: black;
                border-top: 1px solid #ced4da;
            }}
            QCheckBox {{
                font: 9pt "Segoe UI";
                background-color: {self.COLOR_FRAME_BG};
            }}
            QTextEdit {{
                font: 9pt "Segoe UI";
                background-color: white;
                border: 1px solid #ced4da;
            }}
        """

    def create_menu_bar(self):
        print("create_menu_bar started") # DEBUG
        menubar = QMenuBar(self)
        self.setMenuBar(menubar)

        menu_menu = menubar.addMenu("Меню")

        action_reset = QAction("Начать заново", self)
        action_reset.setShortcut("Ctrl+R")
        action_reset.triggered.connect(self.reset_application)
        menu_menu.addAction(action_reset)

        menu_menu.addSeparator()

        action_save_excel = QAction("Сохранить в Excel...", self)
        action_save_excel.setShortcut("Ctrl+S")
        action_save_excel.triggered.connect(self.save_to_excel)
        menu_menu.addAction(action_save_excel)

        import_export_menu = menu_menu.addMenu("Импорт/Экспорт")

        action_save_csv = QAction("Сохранить...", import_export_menu)
        action_save_csv.triggered.connect(self.save_to_csv)
        import_export_menu.addAction(action_save_csv)

        action_load_csv = QAction("Загрузить...", import_export_menu)
        action_load_csv.triggered.connect(self.load_from_csv)
        import_export_menu.addAction(action_load_csv)

        menu_menu.addSeparator()

        action_settings = QAction("Настройки...", menu_menu)
        action_settings.triggered.connect(self.show_settings_dialog)
        menu_menu.addAction(action_settings)
        menu_menu.addSeparator()

        action_about = QAction("О программе...", menu_menu)
        action_about.triggered.connect(self.show_about_window)
        menu_menu.addAction(action_about)
        menu_menu.addSeparator()

        action_debug_console = QAction("Debug Console", menu_menu)
        action_debug_console.triggered.connect(self.create_debug_console)
        menu_menu.addAction(action_debug_console)
        menu_menu.addSeparator()

        action_exit = QAction("Закрыть", self)
        action_exit.setShortcut("Ctrl+Q")
        action_exit.triggered.connect(self.on_closing)
        menu_menu.addAction(action_exit)
        print("create_menu_bar finished") # DEBUG

    def show_settings_dialog(self):
        print("show_settings_dialog started") # DEBUG
        settings_dialog = QDialog(self)
        settings_dialog.setWindowTitle("Настройки")
        settings_layout = QVBoxLayout()
        settings_dialog.setLayout(settings_layout)

        self.strict_validation_checkbox = QCheckBox("Строгая валидация штрихкода")
        self.strict_validation_checkbox.setChecked(self.strict_validation_enabled)
        settings_layout.addWidget(self.strict_validation_checkbox)

        save_button = QPushButton("Сохранить")
        save_button.clicked.connect(lambda: self.save_settings(settings_dialog))
        settings_layout.addWidget(save_button)

        settings_dialog.exec_()
        print("show_settings_dialog finished") # DEBUG

    def save_settings(self, settings_dialog):
        print("save_settings started") # DEBUG
        self.strict_validation_enabled = self.strict_validation_checkbox.isChecked()
        self.save_state()
        settings_dialog.close()
        print("save_settings finished") # DEBUG

    def create_search_frame(self):
        print("create_search_frame started") # DEBUG
        self.search_frame = QWidget()
        self.main_layout.addWidget(self.search_frame)
        search_layout = QHBoxLayout()
        search_layout.setContentsMargins(5,5,5,0)
        self.search_frame.setLayout(search_layout)

        self.search_label = QLabel("Поиск:")
        search_layout.addWidget(self.search_label)

        self.search_entry = QLineEdit()
        search_layout.addWidget(self.search_entry)
        self.search_entry.textChanged.connect(self.filter_items)
        self.search_entry.setContextMenuPolicy(Qt.CustomContextMenu)
        self.search_entry.customContextMenuRequested.connect(lambda event: self.show_paste_menu(event, self.search_entry))
        self.search_entry.setMaximumWidth(300)

        tooltip_search_entry = ToolTip(self.search_entry)
        tooltip_search_entry.setToolTip("Введите текст для фильтрации списка товаров по штрихкоду короба или товара")

        spacer = QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)
        search_layout.addItem(spacer)
        print("create_search_frame finished") # DEBUG

    def create_box_frame(self):
        print("create_box_frame started") # DEBUG
        self.box_frame = QGroupBox("Короб")
        self.main_layout.addWidget(self.box_frame)
        box_layout = QGridLayout()
        self.box_frame.setLayout(box_layout)

        self.box_label = QLabel("Штрихкод короба:")
        box_layout.addWidget(self.box_label, 0, 0, 1, 1, Qt.AlignLeft)

        self.box_entry = QLineEdit()
        box_layout.addWidget(self.box_entry, 0, 1, 1, 1)
        self.box_entry.setMinimumWidth(200)
        self.box_entry.returnPressed.connect(self.process_box_barcode)
        self.box_entry.setContextMenuPolicy(Qt.CustomContextMenu)
        self.box_entry.customContextMenuRequested.connect(lambda event: self.show_paste_menu(event, self.box_entry))
        self.box_entry.setFocus()

        self.new_box_button = QPushButton("Новый короб")
        box_layout.addWidget(self.new_box_button, 0, 2, 1, 1)
        self.new_box_button.clicked.connect(self.new_box)

        tooltip_new_box = ToolTip(self.new_box_button)
        tooltip_new_box.setToolTip("Начать работу с новым коробом (Ctrl+N)")

        tooltip_box_entry = ToolTip(self.box_entry)
        tooltip_box_entry.setToolTip("Введите или отсканируйте штрихкод короба")

        box_layout.setColumnStretch(1, 1)
        print("create_box_frame finished") # DEBUG

    def create_item_scan_frame(self):
        print("create_item_scan_frame started") # DEBUG
        self.item_scan_frame = QGroupBox("Сканирование товаров")
        self.main_layout.addWidget(self.item_scan_frame)
        item_scan_layout = QGridLayout()
        self.item_scan_frame.setLayout(item_scan_layout)

        self.item_scan_label = QLabel("Штрихкод товара:")
        item_scan_layout.addWidget(self.item_scan_label, 0, 0, 1, 1, Qt.AlignLeft)

        self.item_scan_entry = QLineEdit()
        item_scan_layout.addWidget(self.item_scan_entry, 0, 1, 1, 1)
        self.item_scan_entry.setMinimumWidth(200)
        self.item_scan_entry.returnPressed.connect(self.process_item_barcode)
        self.item_scan_entry.setContextMenuPolicy(Qt.CustomContextMenu)
        self.item_scan_entry.customContextMenuRequested.connect(lambda event: self.show_paste_menu(event, self.item_scan_entry))
        self.item_scan_entry.setEnabled(False)

        tooltip_item_entry = ToolTip(self.item_scan_entry)
        tooltip_item_entry.setToolTip("Введите или отсканируйте штрихкод товара")

        item_scan_layout.setColumnStretch(1, 1)

        self.autoclear_item_entry = QCheckBox("Очищать поле ввода")
        self.autoclear_item_entry.setChecked(True)
        item_scan_layout.addWidget(self.autoclear_item_entry, 0, 2, 1, 1, Qt.AlignLeft)

        tooltip_autoclear = ToolTip(self.autoclear_item_entry)
        tooltip_autoclear.setToolTip("Автоматически очищать поле ввода штрихкода товара после каждого сканирования")
        print("create_item_scan_frame finished") # DEBUG

    def create_items_frame(self):
        print("create_items_frame started") # DEBUG
        self.items_frame = QGroupBox("Товары")
        self.main_layout.addWidget(self.items_frame)
        items_layout = QVBoxLayout()
        self.items_frame.setLayout(items_layout)

        self.items_tree = QTreeWidget()
        items_layout.addWidget(self.items_tree)
        self.items_tree.setColumnCount(4)
        self.items_tree.setHeaderLabels(["Штрихкод короба", "Штрихкод товара", "Количество", "Комментарий"])
        self.items_tree.header().setSectionResizeMode(QHeaderView.Stretch)
        self.items_tree.header().setSectionResizeMode(0, QHeaderView.Interactive)
        self.items_tree.header().setSectionResizeMode(1, QHeaderView.Interactive)
        self.items_tree.header().setSectionResizeMode(2, QHeaderView.Interactive)
        self.items_tree.header().setSectionResizeMode(3, QHeaderView.Interactive)
        self.items_tree.setColumnWidth(0, 150)
        self.items_tree.setColumnWidth(1, 150)
        self.items_tree.setColumnWidth(2, 80)
        self.items_tree.setAlternatingRowColors(False)
        self.items_tree.itemClicked.connect(self.clear_selection)
        self.items_tree.customContextMenuRequested.connect(self.show_context_menu)
        self.items_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.items_tree.itemDoubleClicked.connect(self.on_double_click)
        for i in range(self.items_tree.columnCount()):
            self.items_tree.headerItem().setTextAlignment(i, Qt.AlignCenter)
        self.items_tree.setStyleSheet("QTreeView::item { text-align: center; }")
        print("create_items_frame finished") # DEBUG

    def create_debug_console(self):
        print("create_debug_console started") # DEBUG
        self.debug_window = QDialog(self)
        self.debug_window.setWindowTitle("Debug Console")
        self.debug_window.setGeometry(100, 100, 600, 300)

        self.debug_text = QTextEdit(self.debug_window)
        layout = QVBoxLayout(self.debug_window)
        layout.addWidget(self.debug_text)
        self.debug_window.setLayout(layout)
        self.debug_window.show()

        sys.stdout = self
        sys.stderr = self
        print("create_debug_console finished") # DEBUG

    def write(self, message):
        self.debug_text.moveCursor(self.debug_text.textCursor().End)
        self.debug_text.insertPlainText(message)

    def flush(self):
        pass

    def create_control_frame(self):
        print("create_control_frame started") # DEBUG
        self.control_frame = QWidget()
        self.main_layout.addWidget(self.control_frame)
        control_layout = QHBoxLayout()
        self.control_frame.setLayout(control_layout)

        self.history_button = QPushButton("История")
        control_layout.addWidget(self.history_button)
        self.history_button.clicked.connect(self.show_history)

        tooltip_history_button = ToolTip(self.history_button)
        tooltip_history_button.setToolTip("Открыть окно истории сканирования")

        self.summary_label = QLabel("")
        control_layout.addWidget(self.summary_label)

        self.save_button = QPushButton("Сохранить в Excel")
        control_layout.addWidget(self.save_button)
        self.save_button.clicked.connect(self.save_to_excel)
        self.save_button.setEnabled(False)

        tooltip_save_button = ToolTip(self.save_button)
        tooltip_save_button.setToolTip("Сохранить данные в файл Excel (Ctrl+S)")

        control_layout.addStretch()
        print("create_control_frame finished") # DEBUG

    def create_status_bar(self):
        print("create_status_bar started") # DEBUG
        self.status_bar = self.statusBar()
        self.status_bar.setStyleSheet(f"QStatusBar{{background-color: {self.COLOR_HEADER_BG}; border-top: 1px solid #ced4da;}}")
        print("create_status_bar finished") # DEBUG

    def convert_ru_to_en_layout_box(self, barcode):
        print(f"convert_ru_to_en_layout_box started with barcode: {barcode}") # DEBUG
        if barcode.lower().startswith('ца'):
            barcode = 'wb' + barcode[2:]
            print(f"Converted box barcode to: {barcode}") # DEBUG
        elif barcode.lower().startswith('ци_'):  # Recognize and convert "ЦИ_" to "WB_"
            barcode = 'WB_' + barcode[3:] # Keep "WB_" uppercase
            print(f"Converted box barcode ЦИ_ to: {barcode}") # DEBUG
        return barcode

    def convert_ru_to_en_layout_item(self, barcode):
        print(f"convert_ru_to_en_layout_item started with barcode: {barcode}") # DEBUG
        if barcode.lower().startswith('щят'):
            barcode = 'OZN' + barcode[3:]  # Use 'OZN' in uppercase
            print(f"Converted item barcode to: {barcode}") # DEBUG
        return barcode

    def process_box_barcode(self):
        print("process_box_barcode started") # DEBUG
        barcode_input = self.box_entry.text().strip()
        print(f"Input box barcode: {barcode_input}") # DEBUG

        barcode = self.convert_ru_to_en_layout_box(barcode_input) # Auto-convert layout
        if not barcode:
            self.show_warning("Введите штрихкод короба!")
            print("process_box_barcode - Warning: Empty barcode") # DEBUG
            return

        if not self.is_valid_barcode(barcode, barcode_type='box'):
            self.show_error("Неверный штрихкод короба!")
            self.box_entry.clear()
            print("process_box_barcode - Error: Invalid barcode") # DEBUG
            return

        if barcode not in self.all_boxes:
            self.all_boxes[barcode] = {}
            print(f"process_box_barcode - New box added: {barcode}") # DEBUG
        else:
            print(f"process_box_barcode - Existing box: {barcode}") # DEBUG

        self.current_box_barcode = barcode
        self.box_entry.setEnabled(False)
        self.item_scan_entry.setEnabled(True)
        self.item_scan_entry.setFocus()
        self.save_button.setEnabled(True)
        self.update_status(f"Текущий короб: {self.current_box_barcode}")
        self.refresh_treeview()
        self.log_scan(barcode, "box")
        self.highlight_entry(self.box_entry)
        print("process_box_barcode finished") # DEBUG

    def log_scan(self, barcode, barcode_type):
        print(f"log_scan started - type: {barcode_type}, barcode: {barcode}") # DEBUG
        if self.history_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.history_file = os.path.join(self.log_dir, f"scan_history_{timestamp}.log")
            print(f"log_scan - History file created: {self.history_file}") # DEBUG
        try:
            with open(self.history_file, "a") as f:
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                f.write(f"{timestamp} - {barcode_type.upper()}: {barcode}\n")
            print(f"log_scan - Logged: {barcode_type}, {barcode}") # DEBUG
        except Exception as e:
            self.show_error(f"Ошибка при записи в историю: {e}")
            print(f"log_scan - Error writing to history: {e}") # DEBUG
        print("log_scan finished") # DEBUG

    def show_history(self):
        print("show_history started") # DEBUG
        if self.history_window and self.history_window.isVisible():
            self.history_window.raise_()
            self.history_window.activateWindow()
            print("show_history - History window already visible, raised") # DEBUG
            return

        self.history_window = QDialog(self)
        self.history_window.setWindowTitle("История сканирования")
        self.history_window.setGeometry(100, 100, 600, 400)
        layout = QVBoxLayout(self.history_window)

        filter_frame = QWidget()
        layout.addWidget(filter_frame)
        filter_layout = QHBoxLayout(filter_frame)

        filter_label = QLabel("Фильтр истории:")
        filter_layout.addWidget(filter_label)

        self.history_filter_entry = QLineEdit()
        filter_layout.addWidget(self.history_filter_entry)
        self.history_filter_entry.textChanged.connect(self.filter_history)
        self.history_filter_entry.setContextMenuPolicy(Qt.CustomContextMenu)
        self.history_filter_entry.customContextMenuRequested.connect(lambda event: self.show_paste_menu(event, self.history_filter_entry))

        self.history_tree = QTreeWidget()
        layout.addWidget(self.history_tree)
        self.history_tree.setColumnCount(3)
        self.history_tree.setHeaderLabels(["Время", "Тип", "Штрихкод"])
        self.history_tree.header().setSectionResizeMode(QHeaderView.Stretch)
        self.history_tree.header().setSectionResizeMode(0, QHeaderView.Interactive)
        self.history_tree.header().setSectionResizeMode(1, QHeaderView.Interactive)
        self.history_tree.header().setSectionResizeMode(2, QHeaderView.Interactive)
        self.history_tree.setColumnWidth(0, 150)
        self.history_tree.setColumnWidth(1, 50)
        self.history_tree.setColumnWidth(2, 300)

        self.load_history()
        self.history_window.show()
        print("show_history finished") # DEBUG

    def load_history(self):
        print("load_history started") # DEBUG
        self.history_tree.clear()

        if not self.history_file:
            print("load_history - No history file") # DEBUG
            return

        try:
            if os.path.exists(self.history_file):
                print(f"load_history - Loading history from: {self.history_file}") # DEBUG
                with open(self.history_file, "r") as f:
                    for line in f:
                        line = line.strip()
                        if not line:
                            continue

                        try:
                            timestamp_str, rest = line.split(" - ", 1)
                            barcode_type, barcode = rest.split(": ", 1)
                            barcode_type = barcode_type.strip().lower()
                            barcode = barcode.strip()
                            QTreeWidgetItem(self.history_tree, [timestamp_str, barcode_type, barcode])
                        except ValueError:
                            print(f"load_history - Error parsing history line: '{line}'") # DEBUG
                            continue

        except FileNotFoundError:
            print("load_history - File history not found.") # DEBUG
        except Exception as e:
            self.show_error(f"Ошибка при загрузке истории: {e}")
            print(f"load_history - Error loading history: {e}") # DEBUG
        print("load_history finished") # DEBUG

    def filter_history(self):
        print("filter_history started") # DEBUG
        filter_text = self.history_filter_entry.text().lower()
        root = self.history_tree.invisibleRootItem()
        child_count = root.childCount()
        for i in range(child_count):
            item = root.child(i)
            values = [item.text(col) for col in range(self.history_tree.columnCount())]
            if any(filter_text in str(v).lower() for v in values):
                item.setHidden(False)
            else:
                item.setHidden(True)
        print("filter_history finished") # DEBUG

    def create_tooltip(self, widget, text, delay=500):
        tooltip = ToolTip(widget)
        tooltip.delay = delay
        tooltip.setToolTip(text)
        return tooltip

    def show_about_window(self):
        print("show_about_window started") # DEBUG
        try:
            if hasattr(self, 'about_window') and self.about_window and self.about_window.isVisible():
                self.about_window.raise_()
                self.about_window.activateWindow()
                print("show_about_window - About window already visible, raised") # DEBUG
                return

            self.about_window = QDialog(self)
            self.about_window.setWindowTitle("О программе")
            self.about_window.setFixedSize(400, 350)
            self.about_window.setWindowModality(Qt.WindowModal)

            about_layout = QVBoxLayout()
            about_layout.setContentsMargins(20, 20, 20, 20)
            self.about_window.setLayout(about_layout)

            app_name_label = QLabel("ScanBox")
            app_name_label.setFont(QFont("Segoe UI", 12, QFont.Bold))
            app_name_label.setAlignment(Qt.AlignCenter)
            about_layout.addWidget(app_name_label)

            version_label = QLabel("Версия 0.1.0.5 ALPHA Coldfix")
            version_label.setFont(QFont("Segoe UI", 10))
            version_label.setAlignment(Qt.AlignCenter)
            about_layout.addWidget(version_label)

            try:
                base_path = sys._MEIPASS if hasattr(sys, '_MEIPASS') else os.path.dirname(os.path.abspath(__file__))
                image_path = os.path.join(base_path, "resources", "about_image.png")
                if os.path.exists(image_path):
                    about_image = QPixmap(image_path)
                    if not about_image.isNull():
                        image_label = QLabel()
                        image_label.setPixmap(about_image.scaledToWidth(200, Qt.SmoothTransformation))
                        image_label.setAlignment(Qt.AlignCenter)
                        about_layout.addWidget(image_label)
                else:
                    print(f"Не удалось загрузить изображение из пути: {image_path}")

            except Exception as e:
                print(f"Не удалось загрузить изображение 'about_image.png': {e}")

            description_text = "Полностью перенесён функционал с tkinter в  pyqt вариант, будет ли оно работать - хз, всё для Алексея и дальнейших доработок, сырости этой реализации позавидует даже СПб, так что надеемся не отвалится :>"
            description_label = QLabel(description_text)
            description_label.setAlignment(Qt.AlignCenter)
            description_label.setWordWrap(True)
            description_label.setFont(QFont("Segoe UI", 9))
            about_layout.addWidget(description_label)

            copyright_label = QLabel("© 2025, Holorigg")
            copyright_font = QFont("Segoe UI", 9)
            copyright_font.setItalic(True)
            copyright_label.setFont(copyright_font)
            copyright_label.setAlignment(Qt.AlignCenter)
            copyright_label.setStyleSheet("color: #777;")
            about_layout.addWidget(copyright_label)

            ok_button = QPushButton("OK")
            ok_button.setDefault(True)
            ok_button.clicked.connect(self.about_window.close)
            about_layout.addWidget(ok_button)

            self.about_window.show()
            self.about_window.activateWindow()

        except Exception as error:
            self.show_error(f"Ошибка в окне 'О программе': {error}")
            print(f"show_about_window - Error showing about window: {error}") # DEBUG
        print("show_about_window finished") # DEBUG

    def process_item_barcode(self):
        print("process_item_barcode started") # DEBUG
        barcode_input = self.item_scan_entry.text().strip()
        print(f"Input item barcode: {barcode_input}") # DEBUG

        barcode = self.convert_ru_to_en_layout_item(barcode_input) # Auto-convert layout

        if not self.current_box_barcode:
            self.show_warning("Сначала отсканируйте штрихкод короба!")
            self.item_scan_entry.clear()
            self.box_entry.setFocus()
            print("process_item_barcode - Warning: No box barcode scanned first") # DEBUG
            return
        if not barcode:
            self.show_warning("Введите штрихкод товара!")
            print("process_item_barcode - Warning: Empty item barcode") # DEBUG
            return
        if not self.is_valid_barcode(barcode, barcode_type='item'):
            self.show_error("Неверный штрихкод товара!")
            self.item_scan_entry.clear()
            print("process_item_barcode - Error: Invalid item barcode") # DEBUG
            return
        if self.current_box_barcode not in self.all_boxes:
            QMessageBox.showerror(self, "Ошибка", "Текущий короб не найден!")
            print("process_item_barcode - Error: Current box not found in all_boxes") # DEBUG
            return
        self.add_item(barcode)
        self.log_scan(barcode, "item")
        if self.autoclear_item_entry.isChecked():
            self.item_scan_entry.clear()
        self.highlight_entry(self.item_scan_entry)
        self.save_state()
        print("process_item_barcode finished") # DEBUG

    def highlight_entry(self, entry):
        original_bg = entry.styleSheet()
        entry.setStyleSheet("QLineEdit { background-color: #c8e6c9; }")
        QTimer.singleShot(200, lambda: entry.setStyleSheet(""))

    def add_item(self, item_barcode):
        print(f"add_item started with item_barcode: {item_barcode}") # DEBUG
        if item_barcode in self.all_boxes[self.current_box_barcode]:
            self.all_boxes[self.current_box_barcode][item_barcode] += 1
            print(f"add_item - Item count incremented for {item_barcode} in box {self.current_box_barcode}") # DEBUG
        else:
            self.all_boxes[self.current_box_barcode][item_barcode] = 1
            print(f"add_item - New item added {item_barcode} to box {self.current_box_barcode}") # DEBUG
        self.refresh_treeview()
        print("add_item finished") # DEBUG

    def refresh_treeview(self):
        print("refresh_treeview started") # DEBUG
        self.items_tree.clear()
        for box_barcode, items in self.all_boxes.items():
            box_comment = self.comments.get((box_barcode, ""), "")
            box_item = QTreeWidgetItem(self.items_tree, [box_barcode, "", "", box_comment])
            box_item.setFlags(box_item.flags() | Qt.ItemIsTristate)
            box_item.setBackground(0, QColor(self.box_bg_color))
            box_item.setBackground(1, QColor(self.box_bg_color))
            box_item.setBackground(2, QColor(self.box_bg_color))
            box_item.setBackground(3, QColor(self.box_bg_color))
            self.items_tree.expandItem(box_item)

            for item_barcode, count in items.items():
                item_comment = self.comments.get((box_barcode, item_barcode), "")
                if not self.search_query or self.search_query.lower() in box_barcode.lower() or self.search_query.lower() in item_barcode.lower():
                    item = QTreeWidgetItem(box_item, ["", item_barcode, str(count), item_comment])
                    for i in range(1, 4):
                        item.setTextAlignment(i, Qt.AlignCenter)
        self.update_summary()
        print("refresh_treeview finished") # DEBUG

    def filter_items(self):
        self.search_query = self.search_entry.text()
        self.refresh_treeview()

    def show_context_menu(self, point):
        item = self.items_tree.itemAt(point)
        if item is None:
            return

        self.items_tree.setCurrentItem(item)
        column_index = -1

        header = self.items_tree.header()
        x_click = point.x()

        for i in range(self.items_tree.columnCount()):
            section_pos = header.sectionViewportPosition(i)
            section_width = header.sectionSize(i)
            if x_click >= section_pos and x_click < section_pos + section_width:
                column_index = i
                break

        if column_index == -1:
            pass
        else:
            pass

        values = [item.text(i) for i in range(self.items_tree.columnCount())]

        context_menu = QMenu(self)

        if len(values) == 4 and values[1] == "" and values[2] == "":
            if column_index == 0:
                action_copy_box_barcode = QAction("Копировать штрихкод короба", self)
                action_copy_box_barcode.triggered.connect(lambda: self.clipboard.setText(values[0]))
                context_menu.addAction(action_copy_box_barcode)
            elif column_index == 3:
                action_edit_comment = QAction("Изменить комментарий к коробу", self)
                action_edit_comment.triggered.connect(lambda: self.edit_comment(item))
                context_menu.addAction(action_edit_comment)
            else:
                pass

            action_edit_box_barcode = QAction("Изменить штрихкод короба", self)
            action_edit_box_barcode.triggered.connect(lambda: self.edit_box_barcode(item))
            context_menu.addAction(action_edit_box_barcode)

            action_delete_box = QAction("Удалить короб", self)
            action_delete_box.triggered.connect(lambda: self.delete_box(item))
            context_menu.addAction(action_delete_box)

        else:
            parent_item = item.parent()
            box_barcode = parent_item.text(0) if parent_item else ""

            if column_index == 0:
                action_copy_box_barcode = QAction("Копировать штрихкод короба", self)
                action_copy_box_barcode.triggered.connect(lambda: self.clipboard.setText(box_barcode))
                context_menu.addAction(action_copy_box_barcode)
            elif column_index == 1:
                action_copy_item_barcode = QAction("Копировать штрихкод товара", self)
                action_copy_item_barcode.triggered.connect(lambda: self.clipboard.setText(values[1]))
                context_menu.addAction(action_copy_item_barcode)
            elif column_index == 2:
                action_copy_count = QAction("Копировать количество", self)
                action_copy_count.triggered.connect(lambda: self.clipboard.setText(values[2]))
                context_menu.addAction(action_copy_count)
            elif column_index == 3:
                action_edit_comment = QAction("Изменить комментарий к товару", self)
                action_edit_comment.triggered.connect(lambda: self.edit_comment(item))
                context_menu.addAction(action_edit_comment)
            else:
                pass

            if column_index in (1, 2):
                action_edit_count = QAction("Изменить количество", self)
                action_edit_count.triggered.connect(lambda: self.edit_item_count(item))
                context_menu.addAction(action_edit_count)
            if column_index == 1:
                action_edit_item_barcode = QAction('Изменить штрихкод товара', self)
                action_edit_item_barcode.triggered.connect(lambda: self.edit_item_barcode(item))
                context_menu.addAction(action_edit_item_barcode)

            action_delete_item = QAction("Удалить товар", self)
            action_delete_item.triggered.connect(lambda: self.delete_item(item))
            context_menu.addAction(action_delete_item)

        context_menu.popup(self.items_tree.viewport().mapToGlobal(point))

    def clear_selection(self, item, column):
        if not item.isSelected():
            self.items_tree.clearSelection()

    def edit_item_count(self, selected_item):
        parent_item = selected_item.parent()
        box_barcode = parent_item.text(0) if parent_item else ""
        current_count = selected_item.text(2)
        barcode = selected_item.text(1)

        new_count, ok = QInputDialog.getInt(self, "Изменить количество",
                                             f"Введите новое количество для {barcode}:",
                                             int(current_count), 0)
        if ok:
            if str(box_barcode) in self.all_boxes:
                if new_count == 0:
                    if barcode in self.all_boxes[box_barcode]:
                        del self.all_boxes[box_barcode][barcode]
                        if not self.all_boxes[box_barcode]:
                            del self.all_boxes[box_barcode]
                else:
                    self.all_boxes[str(box_barcode)][barcode] = new_count
            selected_item.setText(2, str(new_count))
            self.refresh_treeview()
            self.update_summary()
            self.save_state()

    def edit_box_barcode(self, item):
        old_barcode = item.text(0)

        new_barcode, ok = QInputDialog.getText(self, "Изменить штрихкод короба",
                                            "Введите новый штрихкод короба:",
                                            QLineEdit.Normal, old_barcode)
        if ok and new_barcode and new_barcode != old_barcode:
            if self.is_valid_barcode(new_barcode, barcode_type='box'):
                if new_barcode not in self.all_boxes:
                    self.all_boxes[new_barcode] = self.all_boxes.pop(old_barcode)
                    for key in list(self.comments.keys()):
                        if key[0] == old_barcode:
                            new_key = (new_barcode, key[1])
                            self.comments[new_key] = self.comments.pop(key)

                    if self.current_box_barcode == old_barcode:
                        self.current_box_barcode = new_barcode
                        self.update_status(f"Текущий короб: {self.current_box_barcode}")
                    self.refresh_treeview()
                else:
                    self.show_error("Короб с таким штрихкодом уже существует!")
            else:
                self.show_error("Неверный штрихкод короба!")

    def edit_item_barcode(self, item):
        parent_item = item.parent()
        box_barcode = parent_item.text(0) if parent_item else ""
        old_barcode = item.text(1)

        new_barcode, ok = QInputDialog.getText(self, "Изменить штрихкод товара",
                                            "Введите новый штрихкод товара:",
                                            QLineEdit.Normal, old_barcode)
        if ok and new_barcode and new_barcode != old_barcode:
            if self.is_valid_barcode(new_barcode, barcode_type='item'):
                if new_barcode not in self.all_boxes[box_barcode]:
                    self.all_boxes[box_barcode][new_barcode] = self.all_boxes[box_barcode].pop(old_barcode)
                    if (box_barcode, old_barcode) in self.comments:
                        self.comments[(box_barcode, new_barcode)] = self.comments.pop((box_barcode, old_barcode))
                    self.refresh_treeview()
                else:
                    self.show_error("Товар с таким штрихкодом уже есть в этом коробе!")
            else:
                self.show_error("Неверный штрихкод товара!")

    def delete_box(self, item):
        box_barcode = item.text(0)
        if QMessageBox.question(self, "Удалить короб", f"Вы уверены, что хотите удалить короб '{box_barcode}'?",
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            del self.all_boxes[box_barcode]
            keys_to_delete = []
            for key in self.comments:
                if key[0] == box_barcode:
                    keys_to_delete.append(key)
            for key in keys_to_delete:
                del self.comments[key]
            if self.current_box_barcode == box_barcode:
                self.current_box_barcode = ""
                self.update_status("")
            self.refresh_treeview()

    def delete_item(self, item):
        parent_item = item.parent()
        box_barcode = parent_item.text(0) if parent_item else ""
        item_barcode = item.text(1)

        if QMessageBox.question(self, "Удалить товар", f"Вы уверены, что хотите удалить товар '{item_barcode}' из короба '{box_barcode}'?",
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            del self.all_boxes[box_barcode][item_barcode]
            if (box_barcode, item_barcode) in self.comments:
                del self.comments[(box_barcode, item_barcode)]
            if not self.all_boxes[box_barcode]:
                del self.all_boxes[box_barcode]
            if (box_barcode, "") in self.comments:
                del self.comments[(box_barcode, "")]
            if self.current_box_barcode == box_barcode:
                self.current_box_barcode = ""
                self.update_status("")
            self.refresh_treeview()

    def edit_comment(self, item):
        values = [item.text(i) for i in range(self.items_tree.columnCount())]
        if len(values) == 4 and values[1] == "" and values[2] == "":
            box_barcode = values[0]
            current_comment = self.comments.get((box_barcode, ""), "")
            new_comment, ok = QInputDialog.getText(self, "Изменить комментарий",
                                                f"Введите комментарий для короба {box_barcode}:",
                                                QLineEdit.Normal, current_comment)
            if ok:
                self.comments[(box_barcode, "")] = new_comment
                self.refresh_treeview()
        else:
            parent_item = item.parent()
            box_barcode = parent_item.text(0) if parent_item else ""
            item_barcode = values[1]
            current_comment = self.comments.get((box_barcode, item_barcode), "")
            new_comment, ok = QInputDialog.getText(self, "Изменить комментарий",
                                                f"Введите комментарий для товара {item_barcode}:",
                                                QLineEdit.Normal, current_comment)
            if ok:
                self.comments[(box_barcode, item_barcode)] = new_comment
                self.refresh_treeview()

    def on_double_click(self, item, column_index):
        if column_index in [2]:
            self.edit_item_count(item)

    def save_to_excel(self):
        if not self.all_boxes:
            self.show_warning("Нет данных для сохранения!")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить в Excel", "", "Excel Files (*.xlsx);;All Files (*)")
        if not file_path:
            return
        if not file_path.lower().endswith(('.xlsx')):
            file_path += '.xlsx'
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            for box_barcode, items in self.all_boxes.items():
                sheet = wb.create_sheet(title=f"Короб {box_barcode}")
                sheet['A1'] = "Штрихкод короба"
                sheet['B1'] = box_barcode
                sheet['C1'] = "Комментарий"
                sheet['A2'] = "Штрихкод товара"
                sheet['B2'] = "Количество"
                sheet['C2'] = "Комментарий"
                for cell in ['A1', 'B1', 'C1', 'A2', 'B2', 'C2']:
                    sheet[cell].alignment = Alignment(horizontal='center')
                row = 3
                sheet.cell(row=row, column=1, value="Комментарий к коробу:")
                sheet.cell(row=row, column=3, value=self.comments.get((box_barcode, ""), ""))
                row += 1
                for item_barcode, count in items.items():
                    sheet.cell(row=row, column=1, value=item_barcode)
                    sheet.cell(row=row, column=2, value=count).alignment = Alignment(horizontal='center')
                    sheet.cell(row=row, column=3,
                                  value=self.comments.get((box_barcode, item_barcode), ""))
                    row += 1

                for column in sheet.columns:
                    max_length = 0
                    col_letter = openpyxl.utils.get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:
                            pass
                    sheet.column_dimensions[col_letter].width = max_length + 2
            wb.save(file_path)
            self.show_info(f"Данные сохранены в {file_path}")
        except Exception as e:
            self.show_error(f"Ошибка при сохранении: {e}")

    def save_to_csv(self):
        if not self.all_boxes:
           self.show_warning("Нет данных для сохранения!")
           return
        file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить в CSV", "", "CSV Files (*.csv);;All Files (*)")
        if not file_path:
            return
        if not file_path.lower().endswith(('.csv')):
            file_path += '.csv'

        try:
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Штрихкод короба", "Комментарий короба", "Штрихкод товара", "Количество", "Комментарий товара"])

                for box_barcode, items in self.all_boxes.items():
                    box_comment = self.comments.get((box_barcode, ""), "")
                    for item_barcode, count in items.items():
                        item_comment = self.comments.get((box_barcode, item_barcode),"")
                        writer.writerow([box_barcode, box_comment, item_barcode, count, item_comment])

            self.show_info(f"Данные сохранены в {file_path}")
        except Exception as e:
            self.show_error(f"Ошибка при сохранении: {e}")

    def load_from_csv(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Загрузить из CSV", "", "CSV Files (*.csv);;All Files (*)")
        if not file_path:
            return

        try:
            with open(file_path, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                header = next(reader, None)
                if header is None:
                    self.show_warning("Файл пуст.")
                    return

                if not (len(header) >= 3 and header[0] == "Штрихкод короба" and header[2] == "Штрихкод товара" and header[3] == "Количество"):
                    self.show_warning("Некорректный формат файла CSV. Ожидаются колонки: Штрихкод короба, Штрихкод товара, Количество")
                    return

                self.all_boxes = {}
                self.current_box_barcode = ""
                self.comments = {}
                for row in reader:
                    if len(row) < 3:
                        self.show_warning(f"Некорректное количество столбцов в строке: {row}")
                        continue

                    box_barcode = row[0].strip()
                    item_barcode = row[2].strip()
                    count_str = row[3].strip()

                    box_comment = row[1].strip() if len(row) > 1 else ""
                    item_comment = row[4].strip() if len(row) > 4 else ""

                    if not self.is_valid_barcode(box_barcode, barcode_type='box'):
                        self.show_warning(f'Недопустимый штрихкод короба: {box_barcode}')
                        continue
                    if not self.is_valid_barcode(item_barcode, barcode_type='item'):
                        self.show_warning(f'Недопустимый штрихкод товара: {item_barcode}')
                        continue
                    try:
                        count = int(count_str)
                        if count <= 0:
                            raise ValueError("Количество должно быть положительным")
                    except ValueError:
                        self.show_warning(f"Некорректное количество '{count_str}' для товара '{item_barcode}' в коробе '{box_barcode}'.")
                        continue

                    if box_barcode not in self.all_boxes:
                        self.all_boxes[box_barcode] = {}
                    self.all_boxes[box_barcode][item_barcode] = self.all_boxes[box_barcode].setdefault(item_barcode, 0) + count

                    box_comment = row[1].strip() if len(row) > 1 else ""
                    item_comment = row[4].strip() if len(row) > 4 else ""

                    self.comments[(box_barcode, "")] = box_comment
                    if item_barcode:
                        self.comments[(box_barcode, item_barcode)] = item_comment

                self.refresh_treeview()
                if self.all_boxes:
                    self.update_status("Данные загружены из CSV")
                    self.save_button.setEnabled(True)
        except FileNotFoundError:
            self.show_error("Файл не найден.")
        except Exception as e:
            self.show_error(f"Ошибка при загрузке данных из CSV: {e}")

    def new_box(self):
        print("new_box started") # DEBUG
        self.current_box_barcode = ""
        self.update_status("Введите штрихкод нового короба")
        self.box_entry.setEnabled(True)
        self.box_entry.clear()
        self.box_entry.setFocus()
        self.item_scan_entry.clear()
        self.item_scan_entry.setEnabled(False)
        print("new_box finished") # DEBUG

    def reset_application(self):
        print("reset_application started") # DEBUG
        if QMessageBox.question(self, "Подтверждение", "Вы уверены, что хотите начать заново? Все несохранённые данные будут потеряны.",
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            self.all_boxes = {}
            self.current_box_barcode = ""
            self.search_query = ""
            self.comments = {}
            self.box_entry.setEnabled(True)
            self.box_entry.clear()
            self.item_scan_entry.setEnabled(False)
            self.item_scan_entry.clear()
            self.search_entry.clear()
            self.refresh_treeview()
            self.update_status("")
            self.box_entry.setFocus()
            self.save_button.setEnabled(False)
            self.save_state()
            print("reset_application - Application reset") # DEBUG
        else:
            print("reset_application - Reset cancelled by user") # DEBUG
        print("reset_application finished") # DEBUG

    def is_valid_barcode(self, barcode, barcode_type):
        print(f"is_valid_barcode started - barcode: {barcode}, type: {barcode_type}")  # DEBUG
        if not self.strict_validation_enabled:
            pattern = r"^[\w\-\./]+$"
            valid = bool(re.match(pattern, barcode)) and 8 <= len(barcode) <= 40
            print(f"is_valid_barcode - Validation (old): {valid}")  # DEBUG
            return valid
        else:
            if barcode_type == 'box':
                if barcode.upper().startswith('WB_'):  # Exception for WB_ boxes
                    pattern = r"^WB_[\w\-]+$"  # Stricter pattern for WB_ boxes: WB_ followed by alphanum and hyphen
                    valid = bool(re.match(pattern, barcode.upper())) and 8 <= len(barcode) <= 40  # Keep length check
                    print(f"is_valid_barcode - Validation (strict WB_ exception): {valid}")  # DEBUG
                    return valid
                else:
                    pattern = r"^[0-9]+$"  # Strict pattern for other box codes: only digits
                    valid = bool(re.match(pattern, barcode)) and 8 <= len(barcode) <= 40
                    print(f"is_valid_barcode - Validation (strict box, digits only): {valid}")  # DEBUG
                    return valid
            elif barcode_type == 'item':
                ean13_pattern = r"^[0-9]{13}$" # EAN-13 pattern: 13 digits
                ozn_pattern = r"^ozn[0-9]+$"  # Strict pattern for item codes: ozn followed by digits
                valid_ean13 = bool(re.match(ean13_pattern, barcode))
                valid_ozn = bool(re.match(ozn_pattern, barcode.lower()))

                valid = (valid_ean13 or valid_ozn) and 8 <= len(barcode) <= 40 # Check length only once
                print(f"is_valid_barcode - Validation (strict item, EAN13 or ozn digits): {valid}, EAN13: {valid_ean13}, ozn: {valid_ozn}")  # DEBUG
                return valid
            else:
                print(f"is_valid_barcode - Warning: Unknown barcode type: {barcode_type}. Using default strict (digits only).") #DEBUG
                pattern = r"^[0-9]+$" # Default strict pattern if type is unknown
                valid = bool(re.match(pattern, barcode)) and 8 <= len(barcode) <= 40
                print(f"is_valid_barcode - Validation (strict default, digits only): {valid}")  # DEBUG
                return valid

    def show_error(self, message):
        QMessageBox.critical(self, "Ошибка", message)
        print(f"show_error - Message: {message}") # DEBUG

    def show_warning(self, message):
        QMessageBox.warning(self, "Предупреждение", message)
        print(f"show_warning - Message: {message}") # DEBUG

    def show_info(self, message):
        QMessageBox.information(self, "Информация", message)
        print(f"show_info - Message: {message}") # DEBUG

    def update_status(self, message):
        self.status_bar.showMessage(message)
        print(f"update_status - Message: {message}") # DEBUG

    def update_summary(self):
        num_boxes = len(self.all_boxes)
        total_items = 0
        for box, items in self.all_boxes.items():
            total_items += sum(items.values())

        summary_text = f"Коробов: {num_boxes} | Товаров: {total_items}"
        self.summary_label.setText(summary_text)
        print(f"update_summary - Summary: {summary_text}") # DEBUG

    def load_state(self):
        print("load_state started") # DEBUG
        try:
            if os.path.exists(self.state_file):
                print(f"load_state - Loading state from: {self.state_file}") # DEBUG
                with open(self.state_file, "r") as f:
                    data = json.load(f)
                    if 'all_boxes' in data:
                        self.all_boxes = {str(k): v for k, v in data['all_boxes'].items()}
                    if 'current_box_barcode' in data:
                        self.current_box_barcode = data['current_box_barcode']
                    if 'search_query' in data:
                        self.search_query = data['search_query']
                    serializable_comments = data.get('comments', {})
                    self.comments = {}
                    for key_str, comment in serializable_comments.items():
                        try:
                            box_barcode, item_barcode_str = key_str.split(",", 1) if "," in key_str else (key_str, "")
                            item_barcode = item_barcode_str if item_barcode_str else ""
                            self.comments[(box_barcode, item_barcode)] = comment
                        except ValueError:
                            print(f"load_state - Warning: could not parse comment key string: {key_str}")
                    if 'strict_validation_enabled' in data:
                        self.strict_validation_enabled = data['strict_validation_enabled']
                        if hasattr(self, 'strict_validation_checkbox'):
                            self.strict_validation_checkbox.setChecked(self.strict_validation_enabled)
                print("load_state - State loaded successfully") # DEBUG
                self.refresh_treeview()
                if self.current_box_barcode:
                    self.box_entry.setEnabled(False)
                    self.item_scan_entry.setEnabled(True)
                    self.save_button.setEnabled(True)
        except FileNotFoundError:
            print("load_state - State file not found. Starting fresh.") # DEBUG
        except json.JSONDecodeError as e:
            self.show_error("Ошибка при загрузке состояния: Некорректный формат файла.")
            print(f"load_state - JSONDecodeError: {e}") # DEBUG
        except Exception as e:
            self.show_error(f"Ошибка при загрузке состояния: {e}")
            print(f"load_state - Error loading state: {e}") # DEBUG
        print("load_state finished") # DEBUG

    def save_state(self):
        print("save_state started") # DEBUG
        serializable_comments = {}
        for key, comment in self.comments.items():
            if not isinstance(key, tuple) or len(key) != 2:
                print(f"save_state - WARNING: Invalid key format in self.comments: {key}")
                continue

            box_barcode, item_barcode = key
            key_str = f"{box_barcode},{item_barcode}"
            serializable_comments[key_str] = comment
        data = {
            "all_boxes": self.all_boxes,
            "current_box_barcode": self.current_box_barcode,
            "search_query": self.search_query,
            "comments": serializable_comments,
            "strict_validation_enabled": self.strict_validation_enabled,
        }
        try:
            with open(self.state_file, "w") as f:
                json.dump(data, f)
            print("save_state - State saved successfully") # DEBUG
        except Exception as e:
            self.show_error(f"Ошибка при сохранении состояния: {e}")
            print(f"save_state - Error saving state: {e}") # DEBUG
        print("save_state finished") # DEBUG

    def on_closing(self):
        print("on_closing started") # DEBUG
        self.save_state()
        self.close()
        print("on_closing finished") # DEBUG
        super().closeEvent(QCloseEvent())

    def show_paste_menu(self, event, entry_widget):
        context_menu = QMenu(self)
        paste_action = QAction("Вставить", self)
        paste_action.triggered.connect(lambda: self.paste_from_clipboard(entry_widget))
        context_menu.addAction(paste_action)
        context_menu.popup(entry_widget.mapToGlobal(event))

    def paste_from_clipboard(self, entry_widget):
        text = self.clipboard.text()
        entry_widget.insert(text)


from PyQt5.QtCore import QTimer, QEvent

if __name__ == '__main__':
    app = QApplication(sys.argv)
    barcode_app = QBarcodeApp()
    barcode_app.show()
    sys.exit(app.exec_())
