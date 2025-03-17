import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import openpyxl
from openpyxl.styles import Alignment
import pyzbar.pyzbar as pyzbar
import threading
import os, sys
import pyperclip
import re
import json
from pathlib import Path
import csv
from datetime import datetime


class BarcodeApp:
    def __init__(self, master):
        self.master = master
        master.title("ScanBox")
        
        if hasattr(sys, '_MEIPASS'): 
            base_path = sys._MEIPASS 
        else: 
            base_path = Path(__file__).resolve().parent 

        icon_path = Path(base_path) / "icon.ico" 
        if icon_path.exists():
            try:
                master.iconbitmap(icon_path)
            except tk.TclError:
                print("Не удалось установить иконку")

        self.all_boxes = {}
        self.current_box_barcode = ""
        self.search_query = ""
        self.state_file = "barcode_app_state.json"
        self.box_bg_color = "#f2f2f2"
        self.history_file = None
        self.history_window = None
        self.history_tree = None
        self.comments = {}
        self.history_filter_query = tk.StringVar()

        self.style = ttk.Style()
        self.style.theme_use("default")
        self.bg_color = "#f8f9fa"
        self.frame_bg_color = "#ffffff"
        self.entry_bg_color = "#ffffff"
        self.button_bg_color = "#e0f7fa"
        self.button_fg_color = "#495057"
        self.button_active_bg_color = "#b2ebf2"
        self.button_pressed_bg_color = "#80deea"
        self.header_bg_color = "#f8f9fa"
        self.header_fg_color = "#212529"
        self.scrollbar_bg_color = "#f1f3f5"
        self.scrollbar_trough_color = "#ced4da"
        self.scrollbar_arrow_color = "black"
        self.scrollbar_thumb_color = "#949ca1"

        self.font_label = ("Segoe UI", 9)
        self.font_entry = ("Segoe UI", 9)
        self.font_button = ("Segoe UI Semibold", 10)
        self.font_treeview = ("Segoe UI", 9)
        self.font_treeview_heading = ("Segoe UI Semibold", 11)
        self.font_menu = ("Segoe UI", 10)
        self.configure_ttk_styles()

        self.menubar = tk.Menu(master)
        master.config(menu=self.menubar)
        self.file_menu = tk.Menu(self.menubar, tearoff=0, font=self.font_menu)
        self.menubar.add_cascade(label="Меню", menu=self.file_menu)
        self.file_menu.add_command(label="Начать заново", command=self.reset_application, accelerator="Ctrl+R")
        self.file_menu.add_separator()
        self.file_menu.add_command(label="Сохранить в Excel...", command=self.save_to_excel, accelerator="Ctrl+S")
        self.import_export_menu = tk.Menu(self.file_menu, tearoff=0, font=self.font_menu) # Подменю Импорт/Экспорт
        self.file_menu.add_cascade(label="Импорт/Экспорт", menu=self.import_export_menu) # Каскадное меню
        self.import_export_menu.add_command(label="Сохранить...", command=self.save_to_csv)
        self.import_export_menu.add_command(label="Загрузить...", command=self.load_from_csv)
        self.file_menu.add_separator()
        self.file_menu.add_command(label="О программе...", command=self.show_about_window)
        self.file_menu.add_separator()
        self.file_menu.add_command(label="Debug Console", command=self.create_debug_console)
        self.file_menu.add_separator()
        self.file_menu.add_command(label="Закрыть", command=self.on_closing, accelerator="Ctrl+Q")

        self.main_frame = ttk.Frame(master)
        self.main_frame.pack(padx=10, pady=10, fill="both", expand=True)
        self.create_box_frame()
        self.create_item_scan_frame()
        self.create_items_frame()
        self.create_control_frame()
        self.create_status_bar(master)

        self.load_state()
        master.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.master.bind("<Control-s>", lambda event: self.save_to_excel())
        self.master.bind("<Control-n>", lambda event: self.new_box())
        self.master.bind("<Control-r>", lambda event: self.reset_application())
        self.master.bind("<Control-q>", lambda event: self.on_closing)

    def configure_ttk_styles(self):
        self.style.theme_use("default")

        self.bg_color = "#f8f9fa"
        self.frame_bg_color = "#ffffff"
        self.entry_bg_color = "#ffffff"
        self.button_bg_color = "#e0f7fa"
        self.button_fg_color = "#495057"
        self.button_active_bg_color = "#b2ebf2"
        self.button_pressed_bg_color = "#80deea"
        self.header_bg_color = "#f8f9fa"
        self.header_fg_color = "#212529"
        self.scrollbar_bg_color = "#f1f3f5"
        self.scrollbar_trough_color = "#ced4da"
        self.scrollbar_arrow_color = "black"
        self.scrollbar_thumb_color = "#949ca1"

        self.font_label = ("Segoe UI", 9)
        self.font_entry = ("Segoe UI", 9)
        self.font_button = ("Segoe UI Semibold", 10)
        self.font_treeview = ("Segoe UI", 9)
        self.font_treeview_heading = ("Segoe UI Semibold", 11)
        self.font_menu = ("Segoe UI", 10)

        self.style.configure("TLabel", font=self.font_label, background=self.bg_color)
        self.style.configure("TEntry", font=self.font_entry, fieldbackground=self.entry_bg_color, borderwidth=1, relief="solid", bordercolor="#ced4da")
        self.style.configure("TButton", font=self.font_button, background=self.button_bg_color, foreground=self.button_fg_color, borderwidth=0, relief="flat", padding=(10, 6))
        self.style.map("TButton", background=[("active", self.button_active_bg_color), ("pressed", self.button_pressed_bg_color)],
                       foreground=[("active", self.button_fg_color), ("pressed", self.button_fg_color)])
        self.style.configure("TFrame", background=self.frame_bg_color)
        self.style.configure("TLabelframe", background=self.frame_bg_color, borderwidth=1, relief="groove", bordercolor="#ced4da")
        self.style.configure("TLabelframe.Label", font=("Segoe UI", 11, "bold"), background=self.frame_bg_color, foreground=self.header_fg_color)
        self.style.configure("Treeview", font=self.font_treeview, background="white", fieldbackground="white", borderwidth=1, relief="solid", bordercolor="#ced4da", rowheight=25)
        self.style.configure("Treeview.Heading", font=self.font_treeview_heading, background=self.header_bg_color, foreground=self.header_fg_color, borderwidth=0, relief="flat")
        self.style.map("Treeview.Heading", background=[("active", "#e0e0e0"), ("pressed", "#d0d0d0")])
        self.style.map("Treeview", background=[("selected", "#bbdefb")], foreground=[("selected", "black")])
        self.style.configure("TScrollbar", background=self.scrollbar_bg_color, troughcolor=self.scrollbar_trough_color, arrowcolor=self.scrollbar_arrow_color, bordercolor=self.scrollbar_bg_color, darkcolor=self.scrollbar_bg_color, lightcolor=self.scrollbar_bg_color, gripcount=0, borderwidth=0, relief="flat")
        self.style.map("TScrollbar", background=[("active", self.scrollbar_thumb_color)], arrowcolor=[("pressed", self.scrollbar_arrow_color)])
        self.style.layout("Vertical.TScrollbar", [('Vertical.Scrollbar.trough', {'sticky': 'ns', 'children': [('Vertical.Scrollbar.uparrow', {'side': 'top', 'sticky': ''}), ('Vertical.Scrollbar.downarrow', {'side': 'bottom', 'sticky': ''}), ('Vertical.Scrollbar.thumb', {'expand': '1', 'sticky': 'nswe'})]})])
        self.style.layout("Horizontal.TScrollbar", [('Horizontal.Scrollbar.trough', {'sticky': 'ew', 'children': [('Horizontal.Scrollbar.leftarrow', {'side': 'left', 'sticky': ''}), ('Horizontal.Scrollbar.rightarrow', {'side': 'right', 'sticky': ''}), ('Horizontal.Scrollbar.thumb', {'expand': '1', 'sticky': 'ewns'})]})]) # <-- Исправлено тут: 'side': 'right, '  стало  'side': 'right',

    def create_box_frame(self):
        self.box_frame = ttk.LabelFrame(self.main_frame, text="Короб")
        self.box_frame.pack(fill="x")
        self.box_label = ttk.Label(self.box_frame, text="Штрихкод короба:")
        self.box_label.grid(row=0, column=0, sticky="w", padx=5, pady=(5, 3))
        self.box_entry = ttk.Entry(self.box_frame, width=30)
        self.box_entry.grid(row=0, column=1, padx=5, pady=(5, 3), sticky="ew")
        self.box_entry.bind("<Return>", self.process_box_barcode)
        self.box_entry.bind("<Button-3>", lambda event: self.show_paste_menu(event, self.box_entry))
        self.box_entry.focus_set()
        self.new_box_button = ttk.Button(self.box_frame, text="Новый короб", command=self.new_box)
        self.new_box_button.grid(row=0, column=2, padx=5, pady=(5, 3))
        self.create_tooltip(self.new_box_button, "Начать работу с новым коробом (Ctrl+N)") # Tooltip
        self.create_tooltip(self.box_entry, "Введите или отсканируйте штрихкод короба") # Tooltip
        self.box_frame.columnconfigure(1, weight=1)

    def create_item_scan_frame(self):
        self.item_scan_frame = ttk.LabelFrame(self.main_frame, text="Сканирование товаров")
        self.item_scan_frame.pack(fill="x", pady=(5, 10))
        self.item_scan_label = ttk.Label(self.item_scan_frame, text="Штрихкод товара:")
        self.item_scan_label.grid(row=0, column=0, sticky="w", padx=5, pady=(5, 3))
        self.item_scan_entry = ttk.Entry(self.item_scan_frame, width=30)
        self.item_scan_entry.grid(row=0, column=1, padx=5, pady=(5, 3), sticky="ew")
        self.item_scan_entry.bind("<Return>", self.process_item_barcode)
        self.item_scan_entry.bind("<Button-3>", lambda event: self.show_paste_menu(event, self.item_scan_entry))
        self.item_scan_entry.config(state="disabled")
        self.create_tooltip(self.item_scan_entry, "Введите или отсканируйте штрихкод товара")
        self.item_scan_frame.columnconfigure(1, weight=1)
        self.autoclear_item_entry = tk.BooleanVar(value=True)
        self.autoclear_checkbutton = ttk.Checkbutton(self.item_scan_frame, text="Очищать поле ввода", variable=self.autoclear_item_entry)
        self.autoclear_checkbutton.grid(row=0, column=2, padx=5, pady=(5, 3), sticky='w')
        self.create_tooltip(self.autoclear_checkbutton, "Автоматически очищать поле ввода штрихкода товара после каждого сканирования")


    def create_items_frame(self):
        self.items_frame = ttk.LabelFrame(self.main_frame, text="Товары")
        self.items_frame.pack(fill="both", expand=True)
        self.search_frame = ttk.Frame(self.items_frame)
        self.search_frame.grid(row=0, column=0, sticky="ew", padx=5, pady=(5, 3))
        self.search_label = ttk.Label(self.search_frame, text="Поиск:")
        self.search_label.grid(row=0, column=0, sticky="w", padx=(0, 5))
        self.search_entry = ttk.Entry(self.search_frame)
        self.search_entry.grid(row=0, column=1, sticky="ew", padx=(0, 5))
        self.search_entry.bind("<KeyRelease>", self.filter_items)
        self.search_entry.bind("<Button-3>", lambda event: self.show_paste_menu(event, self.search_entry))
        self.create_tooltip(self.search_entry, "Введите текст для фильтрации списка товаров по штрихкоду короба или товара")
        self.search_frame.columnconfigure(1, weight=0)
        self.items_tree = ttk.Treeview(self.items_frame, columns=("box_barcode", "barcode", "count", "comment"), show="headings", style="Treeview")
        self.items_tree.heading("box_barcode", text="Штрихкод короба")
        self.items_tree.heading("barcode", text="Штрихкод товара")
        self.items_tree.heading("count", text="Количество")
        self.items_tree.heading("comment", text="Комментарий")
        self.items_tree.grid(row=1, column=0, sticky="nsew", padx=(5, 0), pady=5)
        self.items_tree.column("box_barcode", width=150, anchor="center")
        self.items_tree.column("barcode", width=150, anchor="center")
        self.items_tree.column("count", width=80, anchor="center")
        self.items_tree.column("comment", width=200, anchor="w")
        self.items_tree.tag_configure("box_row", background=self.box_bg_color)
        self.tree_yscroll = ttk.Scrollbar(self.items_frame, orient="vertical", command=self.items_tree.yview, style="Vertical.TScrollbar")
        self.tree_yscroll.grid(row=1, column=1, sticky="ns", padx=(0, 5), pady=5)
        self.items_tree.configure(yscrollcommand=self.tree_yscroll.set)
        self.tree_xscroll = ttk.Scrollbar(self.items_frame, orient="horizontal", command=self.items_tree.xview, style="Horizontal.TScrollbar")
        self.tree_xscroll.grid(row=2, column=0, sticky="ew", padx=5, pady=(0, 5), columnspan=2)
        self.items_tree.configure(xscrollcommand=self.tree_xscroll.set)
        self.items_tree.bind("<Button-3>", self.show_context_menu)
        self.items_tree.bind("<Double-1>", self.on_double_click)
        self.items_tree.bind("<Button-1>", self.clear_selection)
        self.items_frame.rowconfigure(1, weight=1)
        self.items_frame.columnconfigure(0, weight=1)
        
    def create_debug_console(self):
        self.debug_window = tk.Toplevel(self.master)
        self.debug_window.title("Debug Console")
        self.debug_window.geometry("600x300")

        self.debug_text = tk.Text(self.debug_window, wrap="word", height=15)
        self.debug_text.pack(fill="both", expand=True)

        sys.stdout = self  
        sys.stderr = self

    def write(self, message):
        self.debug_text.insert("end", message)
        self.debug_text.see("end")

    def flush(self):
        pass

    def create_control_frame(self):
        self.control_frame = ttk.Frame(self.main_frame)
        self.control_frame.pack(fill="x", pady=(5, 0))
        self.save_button = ttk.Button(self.control_frame, text="Сохранить в Excel", command=self.save_to_excel)
        self.save_button.pack(side="right", padx=5)
        self.save_button.config(state="disabled")
        self.create_tooltip(self.save_button, "Сохранить данные в файл Excel (Ctrl+S)")
        self.history_button = ttk.Button(self.control_frame, text="История", command=self.show_history)
        self.history_button.pack(side="left", padx=5)
        self.create_tooltip(self.history_button, "Открыть окно истории сканирования")
        self.summary_label = ttk.Label(self.control_frame, text="", font=("Segoe UI", 9), background=self.frame_bg_color)
        self.summary_label.pack(side="left", padx=5)

    def create_status_bar(self, master):
        self.status_bar = ttk.Label(master, text="", relief="sunken", anchor="w", background=self.header_bg_color)
        self.status_bar.pack(side="bottom", fill="x")

    def process_box_barcode(self, event=None):
        barcode = self.box_entry.get().strip()
        if not barcode:
            self.show_warning("Введите штрихкод короба!")
            return
        if not self.is_valid_barcode(barcode):
            self.show_error("Неверный штрихкод короба!")
            self.box_entry.delete(0, tk.END)
            return
        if barcode not in self.all_boxes:
            self.all_boxes[barcode] = {}
        self.current_box_barcode = barcode
        self.box_entry.config(state="disabled")
        self.item_scan_entry.config(state="normal")
        self.item_scan_entry.focus_set()
        self.save_button.config(state="normal")
        self.update_status(f"Текущий короб: {self.current_box_barcode}")
        self.refresh_treeview()
        self.log_scan(barcode, "box")
        self.highlight_entry(self.box_entry)

    def process_item_barcode(self, event=None):
        barcode = self.item_scan_entry.get().strip()
        if not self.current_box_barcode:
            self.show_warning("Сначала отсканируйте штрихкод короба!")
            self.item_scan_entry.delete(0, tk.END)
            self.box_entry.focus_set()
            return
        if not barcode:
            self.show_warning("Введите штрихкод товара!")
            return
        if not self.is_valid_barcode(barcode):
            self.show_error("Неверный штрихкод товара!")
            self.item_scan_entry.delete(0, tk.END)
            return
        if self.current_box_barcode not in self.all_boxes:
            messagebox.showerror("Ошибка", "Текущий короб не найден!")
            return
        self.add_item(barcode)
        self.log_scan(barcode, "item")
        if self.autoclear_item_entry.get():
            self.item_scan_entry.delete(0, tk.END)
        self.highlight_entry(self.item_scan_entry)

    def highlight_entry(self, entry):
        original_bg = entry.cget("background")
        entry.config(background="#c8e6c9")
        self.master.after(200, lambda: entry.config(background=original_bg))

    def add_item(self, item_barcode):
        if item_barcode in self.all_boxes[self.current_box_barcode]:
            self.all_boxes[self.current_box_barcode][item_barcode] += 1
        else:
            self.all_boxes[self.current_box_barcode][item_barcode] = 1
        self.refresh_treeview()

    def refresh_treeview(self):
        for item in self.items_tree.get_children():
            self.items_tree.delete(item)
        for box_barcode, items in self.all_boxes.items():
            box_comment = self.comments.get((box_barcode, ""), "")
            box_item_id = self.items_tree.insert("", "end", values=(box_barcode, "", "", box_comment), open=True, tags=('box_row',))
            for item_barcode, count in items.items():
                item_comment = self.comments.get((box_barcode, item_barcode), "")
                if not self.search_query or self.search_query.lower() in box_barcode.lower() or self.search_query.lower() in item_barcode.lower():
                    self.items_tree.insert(box_item_id, "end", values=("", item_barcode, count, item_comment))
        self.update_summary()
        self.style.map("Treeview", foreground=[('disabled', 'gray30')])

    def filter_items(self, event=None):
        self.search_query = self.search_entry.get()
        self.refresh_treeview()

    def show_context_menu(self, event):
      item_id = self.items_tree.identify_row(event.y)
      if not item_id:
          return

      self.items_tree.selection_set(item_id)
      column_id = self.items_tree.identify_column(event.x)
      values = self.items_tree.item(item_id, "values")

      context_menu = tk.Menu(self.master, tearoff=0, font=self.font_menu)

      if len(values) == 4 and values[1] == "" and values[2] == "":
          if column_id == "#1":
              context_menu.add_command(label="Копировать штрихкод короба", command=lambda: pyperclip.copy(values[0]))
          elif column_id == '#4':
              context_menu.add_command(label="Изменить комментарий", command=lambda: self.edit_comment(item_id))
          context_menu.add_command(label="Изменить штрихкод короба", command=lambda: self.edit_box_barcode(item_id))
          context_menu.add_command(label="Удалить короб", command=lambda: self.delete_box(item_id))
      else: 
          selected_item_id = self.items_tree.selection()[0]
          item_values = self.items_tree.item(selected_item_id, "values")

          parent_item_id = self.items_tree.parent(item_id)
          box_barcode = self.items_tree.item(parent_item_id, "values")[0]
          if column_id == "#1":
              context_menu.add_command(label="Копировать штрихкод короба",
                                       command=lambda: pyperclip.copy(box_barcode))
          elif column_id == "#2":
              context_menu.add_command(label="Копировать штрихкод товара", command=lambda: pyperclip.copy(values[1]))
          elif column_id == "#3":
              context_menu.add_command(label="Копировать количество", command=lambda: pyperclip.copy(values[2]))
          elif column_id == "#4":
              context_menu.add_command(label="Изменить комментарий", command=lambda: self.edit_comment(item_id))

          if column_id in ("#2", "#3"):
              context_menu.add_command(label="Изменить количество",
                                        command=lambda: self.edit_item_count(selected_item_id))
          if column_id == '#2':
              context_menu.add_command(label='Изменить штрихкод товара', command=lambda: self.edit_item_barcode(item_id))
          context_menu.add_command(label="Удалить товар", command=lambda: self.delete_item(item_id))
      context_menu.post(event.x_root, event.y_root)

    def clear_selection(self, event):
      if self.items_tree.identify_row(event.y) == '':
          self.items_tree.selection_remove(self.items_tree.selection())

    def edit_item_count(self, selected_item_id):
        parent_id = self.items_tree.parent(selected_item_id)
        box_barcode = self.items_tree.item(parent_id)['values'][0]
        current_count = self.items_tree.item(selected_item_id, 'values')[2]
        barcode = self.items_tree.item(selected_item_id, 'values')[1]

        new_count = simpledialog.askinteger(
            "Изменить количество",
            f"Введите новое количество для {barcode}:",
            parent=self.master,
            initialvalue=current_count,
            minvalue=0
        )
        if new_count is not None:
            print(f"Тип box_barcode: {type(box_barcode)}, Значение: {box_barcode}")
            print(f"Ключи all_boxes: {list(self.all_boxes.keys())}")
            print(f"Типы ключей в all_boxes: {[type(k) for k in self.all_boxes.keys()]}")
            if str(box_barcode) in self.all_boxes:
                if new_count == 0:
                    if barcode in self.all_boxes[box_barcode]:
                        del self.all_boxes[box_barcode][barcode]
                        if not self.all_boxes[box_barcode]:
                            del self.all_boxes[box_barcode]
                else:
                    self.all_boxes[str(box_barcode)][barcode] = new_count

            item_values = self.items_tree.item(selected_item_id, 'values')
            if len(item_values) == 4:
                new_values = ('', barcode, new_count, item_values[3])
                self.items_tree.item(selected_item_id, values=new_values)
            print(f"Изменение количества: Короб: {box_barcode}, Товар: {barcode}, Новое количество: {new_count}")
            print(f"all_boxes до изменения: {self.all_boxes}")
            self.refresh_treeview()
            print(f"all_boxes после изменения: {self.all_boxes}")
            self.update_summary()
            self.save_state()

    def edit_box_barcode(self, item_id):
      old_barcode = self.items_tree.item(item_id, "values")[0]

      new_barcode = simpledialog.askstring("Изменить штрихкод короба",
                                          "Введите новый штрихкод короба:",
                                          parent=self.master,
                                          initialvalue=old_barcode)

      if new_barcode is not None and new_barcode != old_barcode:
          if self.is_valid_barcode(new_barcode):
              if new_barcode not in self.all_boxes:
                  self.all_boxes[new_barcode] = self.all_boxes.pop(old_barcode)
                  for key in list(self.comments.keys()):
                      if key[0] == old_barcode:
                          new_key = (new_barcode, key[1])
                          self.comments[new_key] = self.comments.pop(key)

                  if self.current_box_barcode == old_barcode:
                      self.current_box_barcode = new_barcode
                      self.update_status(f"Текущий короб: {self.current_box_barcode}")
                  self.refresh_treeview()
              else:
                  self.show_error("Короб с таким штрихкодом уже существует!")

          else:
              self.show_error("Неверный штрихкод короба!")

    def edit_item_barcode(self, item_id):
      parent_item_id = self.items_tree.parent(item_id)
      box_barcode = self.items_tree.item(parent_item_id, "values")[0]
      old_barcode = self.items_tree.item(item_id, "values")[1]

      new_barcode = simpledialog.askstring("Изменить штрихкод товара",
                                          "Введите новый штрихкод товара:",
                                          parent=self.master,
                                          initialvalue=old_barcode)
      if new_barcode is not None and new_barcode != old_barcode:
          if self.is_valid_barcode(new_barcode):
              if new_barcode not in self.all_boxes[box_barcode]:
                  self.all_boxes[box_barcode][new_barcode] = self.all_boxes[box_barcode].pop(old_barcode)
                  if (box_barcode, old_barcode) in self.comments:
                    self.comments[(box_barcode, new_barcode)] = self.comments.pop((box_barcode, old_barcode))
                  self.refresh_treeview()
              else:
                self.show_error("Товар с таким штрихкодом уже есть в этом коробе!")
          else:
            self.show_error("Неверный штрихкод товара!")

    def delete_box(self, item_id):
      box_barcode = self.items_tree.item(item_id, "values")[0]
      if messagebox.askyesno("Удалить короб", f"Вы уверены, что хотите удалить короб '{box_barcode}'?"):
          del self.all_boxes[box_barcode]
          keys_to_delete = []
          for key in self.comments:
            if key[0] == box_barcode:
                keys_to_delete.append(key)
          for key in keys_to_delete:
            del self.comments[key]
          if self.current_box_barcode == box_barcode:
            self.current_box_barcode = ""
            self.update_status("")
          self.refresh_treeview()

    def delete_item(self, item_id):
      parent_item_id = self.items_tree.parent(item_id)
      box_barcode = self.items_tree.item(parent_item_id, "values")[0]
      item_barcode = self.items_tree.item(item_id, "values")[1]

      if messagebox.askyesno("Удалить товар", f"Вы уверены, что хотите удалить товар '{item_barcode}' из короба '{box_barcode}'?"):
          del self.all_boxes[box_barcode][item_barcode]
          if (box_barcode, item_barcode) in self.comments:
            del self.comments[(box_barcode, item_barcode)]
          if not self.all_boxes[box_barcode]:
            del self.all_boxes[box_barcode]
          if (box_barcode, "") in self.comments:
            del self.comments[(box_barcode, "")]
          if self.current_box_barcode == box_barcode:
            self.current_box_barcode = ""
            self.update_status("")

          self.refresh_treeview()

    def edit_comment(self, item_id):
      values = self.items_tree.item(item_id, "values")
      if len(values) == 4 and values[1] == "" and values[2] == "":  
          box_barcode = values[0]
          current_comment = self.comments.get((box_barcode, ""), "")  
          new_comment = simpledialog.askstring("Изменить комментарий",
                                              f"Введите комментарий для короба {box_barcode}:",
                                              parent=self.master,
                                              initialvalue=current_comment)
          if new_comment is not None:
              self.comments[(box_barcode, "")] = new_comment
              self.refresh_treeview()
      else:
          # Это товар
          parent_item_id = self.items_tree.parent(item_id)
          box_barcode = self.items_tree.item(parent_item_id, "values")[0]
          item_barcode = values[1]
          current_comment = self.comments.get((box_barcode, item_barcode), "")
          new_comment = simpledialog.askstring("Изменить комментарий",
                                              f"Введите комментарий для товара {item_barcode}:",
                                              parent=self.master,
                                              initialvalue=current_comment)
          if new_comment is not None:
              self.comments[(box_barcode, item_barcode)] = new_comment
              self.refresh_treeview()
              
    def on_double_click(self, event):
        item_id = self.items_tree.identify_row(event.y) 
        column_id = self.items_tree.identify_column(event.x)
    
        if not item_id:
            return

        column_index = int(column_id[1:]) - 1
        if column_index < 0:
            return

        values = self.items_tree.item(item_id, "values")
        if len(values) <= column_index:
            return

        current_value = values[column_index]

        x, y, width, height = self.items_tree.bbox(item_id, column_id)

        self.edit_entry = ttk.Entry(self.items_tree)
        self.edit_entry.place(x=x, y=y, width=width, height=height)
        self.edit_entry.insert(0, current_value)
        self.edit_entry.focus()

        self.edit_entry.bind("<Return>", lambda e: self.save_edit(item_id, column_index))
        self.edit_entry.bind("<FocusOut>", lambda e: self.save_edit(item_id, column_index))

    def save_to_excel(self):
      if not self.all_boxes:
          self.show_warning("Нет данных для сохранения!")
          return
      file_path = filedialog.asksaveasfilename(
          defaultextension=".xlsx",
          filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
      )
      if not file_path:
          return
      try:
          wb = openpyxl.Workbook()
          wb.remove(wb.active)
          for box_barcode, items in self.all_boxes.items():
              sheet = wb.create_sheet(title=f"Короб {box_barcode}")
              sheet['A1'] = "Штрихкод короба"
              sheet['B1'] = box_barcode
              sheet['C1'] = "Комментарий"
              sheet['A2'] = "Штрихкод товара"
              sheet['B2'] = "Количество"
              sheet['C2'] = "Комментарий"
              for cell in ['A1', 'B1', 'C1', 'A2', 'B2', 'C2']:
                  sheet[cell].alignment = Alignment(horizontal='center')
              row = 3
              sheet.cell(row=row, column=1, value="Комментарий к коробу:")
              sheet.cell(row=row, column=3, value=self.comments.get((box_barcode, ""), ""))
              row += 1
              for item_barcode, count in items.items():
                sheet.cell(row=row, column=1, value=item_barcode)
                sheet.cell(row=row, column=2, value=count).alignment = Alignment(horizontal='center')
                sheet.cell(row=row, column=3,
                              value=self.comments.get((box_barcode, item_barcode), ""))
                row += 1

              for column in sheet.columns:
                max_length = 0
                col_letter = openpyxl.utils.get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                sheet.column_dimensions[col_letter].width = max_length + 2
          wb.save(file_path)
          self.show_info(f"Данные сохранены в {file_path}")
      except Exception as e:
          self.show_error(f"Ошибка при сохранении: {e}")

    def save_to_csv(self):
        if not self.all_boxes:
           self.show_warning("Нет данных для сохранения!")
           return
        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")],
        )
        if not file_path:
            return

        try:
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Штрихкод короба", "Комментарий короба", "Штрихкод товара", "Количество", "Комментарий товара"])

                for box_barcode, items in self.all_boxes.items():
                    box_comment = self.comments.get((box_barcode, ""), "")
                    for item_barcode, count in items.items():
                        item_comment = self.comments.get((box_barcode, item_barcode),"")
                        writer.writerow([box_barcode, box_comment, item_barcode, count, item_comment])

            self.show_info(f"Данные сохранены в {file_path}")
        except Exception as e:
            self.show_error(f"Ошибка при сохранении: {e}")

    def load_from_csv(self):
      file_path = filedialog.askopenfilename(
          filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")],
      )
      if not file_path:
          return

      try:
          with open(file_path, "r", encoding="utf-8") as f:
              reader = csv.reader(f)
              header = next(reader, None)
              if header is None:
                  self.show_warning("Файл пуст.")
                  return

              if not (len(header) >= 3 and header[0] == "Штрихкод короба" and header[2] == "Штрихкод товара" and header[3] == "Количество"):
                  self.show_warning("Некорректный формат файла CSV. Ожидаются колонки: Штрихкод короба, Штрихкод товара, Количество")
                  return


              self.all_boxes = {}
              self.current_box_barcode = ""
              self.comments = {}
              for row in reader:
                  if len(row) < 3:
                      self.show_warning(f"Некорректное количество столбцов в строке: {row}")
                      continue

                  box_barcode = row[0].strip()
                  item_barcode = row[2].strip()
                  count_str = row[3].strip()

                  box_comment = row[1].strip() if len(row) > 1 else ""
                  item_comment = row[4].strip() if len(row) > 4 else ""

                  if not self.is_valid_barcode(box_barcode):
                      self.show_warning(f'Недопустимый штрихкод короба: {box_barcode}')
                      continue
                  if not self.is_valid_barcode(item_barcode):
                      self.show_warning(f'Недопустимый штрихкод товара: {item_barcode}')
                      continue
                  try:
                      count = int(count_str)
                      if count <= 0:
                          raise ValueError("Количество должно быть положительным")
                  except ValueError:
                      self.show_warning(f"Некорректное количество '{count_str}' для товара '{item_barcode}' в коробе '{box_barcode}'.")
                      continue

                  if box_barcode not in self.all_boxes:
                      self.all_boxes[box_barcode] = {}
                  self.all_boxes[box_barcode][item_barcode] = self.all_boxes[box_barcode].setdefault(item_barcode, 0) + count

                  box_comment = row[1].strip() if len(row) > 1 else "" 
                  item_comment = row[4].strip() if len(row) > 4 else "" 


                  self.comments[(box_barcode, "")] = box_comment 
                  if item_barcode: 
                      self.comments[(box_barcode, item_barcode)] = item_comment


              self.refresh_treeview()
              if self.all_boxes:
                  self.update_status("Данные загружены из CSV")
                  self.save_button.config(state='normal')
      except FileNotFoundError:
          self.show_error("Файл не найден.")
      except Exception as e:
          self.show_error(f"Ошибка при загрузке данных из CSV: {e}")

    def new_box(self):
        self.current_box_barcode = ""
        self.update_status("Введите штрихкод нового короба")
        self.box_entry.config(state="normal")
        self.box_entry.delete(0, tk.END)
        self.box_entry.focus_set()
        self.item_scan_entry.delete(0, tk.END)
        self.item_scan_entry.config(state="disabled")

    def reset_application(self):
        if messagebox.askyesno("Подтверждение", "Вы уверены, что хотите начать заново? Все несохранённые данные будут потеряны."):
            self.all_boxes = {}
            self.current_box_barcode = ""
            self.search_query = ""
            self.comments = {}
            self.box_entry.config(state="normal")
            self.box_entry.delete(0, tk.END)
            self.item_scan_entry.config(state="disabled")
            self.item_scan_entry.delete(0, tk.END)
            self.search_entry.delete(0, tk.END)
            self.refresh_treeview()
            self.update_status("")
            self.box_entry.focus_set()
            self.save_button.config(state='disabled')
            self.save_state()

    def is_valid_barcode(self, barcode):
        pattern = r"^[\w\-\./]+$"
        return bool(re.match(pattern, barcode)) and 8 <= len(barcode) <= 40

    def show_error(self, message):
        messagebox.showerror("Ошибка", message)

    def show_warning(self, message):
        messagebox.showwarning("Предупреждение", message)

    def show_info(self, message):
        messagebox.showinfo("Информация", message)

    def update_status(self, message):
        self.status_bar.config(text=message)

    def update_summary(self):
        num_boxes = len(self.all_boxes)
        total_items = 0
        for box, items in self.all_boxes.items():
            total_items += sum(items.values())

        summary_text = f"Коробов: {num_boxes} | Товаров: {total_items}"
        self.summary_label.config(text=summary_text)

    def load_state(self):
        try:
            with open(self.state_file, "r") as f:
                data = json.load(f)
                if 'all_boxes' in data:
                    self.all_boxes = {str(k): v for k, v in data['all_boxes'].items()}
                    self.all_boxes = data['all_boxes']
                if 'current_box_barcode' in data:
                    self.current_box_barcode = data['current_box_barcode']
                if 'search_query' in data:
                    self.search_query = data['search_query']
                serializable_comments = data.get('comments', {})
                self.comments = {}
                for key_str, comment in serializable_comments.items():
                    try:
                        box_barcode, item_barcode_str = key_str.split(",", 1) if "," in key_str else (key_str, "")
                        item_barcode = item_barcode_str if item_barcode_str else ""
                        self.comments[(box_barcode, item_barcode)] = comment
                    except ValueError:
                        print(f"Warning: could not parse comment key string: {key_str}")

            self.refresh_treeview()
            if self.current_box_barcode:
                self.box_entry.config(state='disabled')
                self.item_scan_entry.config(state='normal')
                self.save_button.config(state='normal')
            print("State loaded successfully.")
        except FileNotFoundError:
            print("State file not found. Starting fresh.")
        except json.JSONDecodeError as e:
            self.show_error("Ошибка при загрузке состояния: Некорректный формат файла.")
            print(f"JSONDecodeError details: {e}")
        except Exception as e:
            self.show_error(f"Ошибка при загрузке состояния: {e}")
            print(f"General load_state error details: {e}")

    def save_state(self):
        serializable_comments = {}
        for key, comment in self.comments.items():
            if not isinstance(key, tuple) or len(key) != 2:
                print(f"WARNING: Invalid key format in self.comments: {key}")
                continue

            box_barcode, item_barcode = key
            key_str = f"{box_barcode},{item_barcode}"
            serializable_comments[key_str] = comment
        data = {
            "all_boxes": self.all_boxes,
            "current_box_barcode": self.current_box_barcode,
            "search_query": self.search_query,
            "comments": serializable_comments,
        }
        try:
            with open(self.state_file, "w") as f:
                json.dump(data, f)
        except Exception as e:
            self.show_error(f"Ошибка при сохранении состояния: {e}")
            print(f"Error details: {e}")

    def save_edit(self, item_id, column_index):
        new_value = self.edit_entry.get().strip()
        self.edit_entry.destroy()  
    
        if not new_value:
            return

        values = list(self.items_tree.item(item_id, "values"))
        old_value = values[column_index]

        if new_value == old_value:
            return

        values[column_index] = new_value
        self.items_tree.item(item_id, values=values)

        parent_id = self.items_tree.parent(item_id)
        if parent_id:
            box_barcode = self.items_tree.item(parent_id, "values")[0]
            item_barcode = values[1]  # Штрихкод товара

            if column_index == 2:
                try:
                    new_count = int(new_value)
                    if new_count <= 0:
                        del self.all_boxes[box_barcode][item_barcode]
                    else:
                        self.all_boxes[box_barcode][item_barcode] = new_count
                except ValueError:
                    self.show_error("Количество должно быть числом!")
                    return
        elif column_index == 1:
            if new_value in self.all_boxes[box_barcode]:
                self.show_error("Товар с таким штрихкодом уже есть!")
                return
                self.all_boxes[box_barcode][new_value] = self.all_boxes[box_barcode].pop(old_value)
            elif column_index == 3:
                self.comments[(box_barcode, item_barcode)] = new_value

        else:
            box_barcode = values[0]
            if column_index == 0:
                if new_value in self.all_boxes:
                    self.show_error("Короб с таким штрихкодом уже существует!")
                    return
                self.all_boxes[new_value] = self.all_boxes.pop(old_value)
                self.comments[(new_value, "")] = self.comments.pop((old_value, ""), "")

            elif column_index == 3:
                self.comments[(box_barcode, "")] = new_value

        self.save_state()
        self.refresh_treeview()
        


    def on_closing(self):
        self.save_state()
        self.master.destroy()

    def show_paste_menu(self, event, entry_widget):
        context_menu = tk.Menu(self.master, tearoff=0, font=self.font_menu)
        context_menu.add_command(label="Вставить", command=lambda: self.paste_from_clipboard(entry_widget))
        context_menu.post(event.x_root, event.y_root)

    def paste_from_clipboard(self, entry_widget):
        try:
            text = self.master.clipboard_get()
            entry_widget.insert(tk.INSERT, text)
        except tk.TclError:
            pass

    def log_scan(self, barcode, barcode_type):
        if self.history_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.history_file = f"scan_history_{timestamp}.log"
        try:
            with open(self.history_file, "a") as f:
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                f.write(f"{timestamp} - {barcode_type.upper()}: {barcode}\n")
        except Exception as e:
            self.show_error(f"Ошибка при записи в историю: {e}")

    def show_history(self):
        if self.history_window and self.history_window.winfo_exists():
            self.history_window.lift()
            return

        self.history_window = tk.Toplevel(self.master)
        self.history_window.title("История сканирования")
        self.history_window.geometry("600x400")

        self.history_filter_frame = ttk.Frame(self.history_window) 
        self.history_filter_frame.pack(side="top", fill="x", padx=5, pady=5)

        self.history_filter_label = ttk.Label(self.history_filter_frame, text="Фильтр истории:")
        self.history_filter_label.pack(side="left", padx=(0, 5))

        self.history_filter_entry = ttk.Entry(self.history_filter_frame, textvariable=self.history_filter_query)
        self.history_filter_entry.pack(side="left", fill="x", expand=True)
        self.history_filter_entry.bind("<KeyRelease>", self.filter_history)
        self.history_filter_entry.bind("<Button-3>", lambda event: self.show_paste_menu(event, self.history_filter_entry))

        self.history_tree = ttk.Treeview(self.history_window, columns=("timestamp", "type", "barcode"), show="headings")
        self.history_tree.heading("timestamp", text="Время")
        self.history_tree.heading("type", text="Тип")
        self.history_tree.heading("barcode", text="Штрихкод")
        self.history_tree.pack(fill="both", expand=True, padx=5, pady=(0, 5))

        self.history_tree.column("timestamp", width=150, anchor="center")
        self.history_tree.column("type", width=50, anchor="center")
        self.history_tree.column("barcode", width=300, anchor="center")

        self.load_history()

    def load_history(self):
        for item in self.history_tree.get_children():
            self.history_tree.delete(item)

        if not self.history_file:
            return

        try:
            with open(self.history_file, "r") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue

                    try:
                        timestamp_str, rest = line.split(" - ", 1)
                        barcode_type, barcode = rest.split(": ", 1)
                        barcode_type = barcode_type.strip().lower()
                        barcode = barcode.strip()
                        self.history_tree.insert("", "end", values=(timestamp_str, barcode_type, barcode))

                    except ValueError:
                        print(f"Ошибка парсинга строки в истории: '{line}'")
                        continue

        except FileNotFoundError:
            print("Файл истории не найден.")
        except Exception as e:
            self.show_error(f"Ошибка при загрузке истории: {e}")

    def filter_history(self, event=None):
        filter_text = self.history_filter_query.get().lower()
        for item in self.history_tree.get_children():
            values = self.history_tree.item(item, 'values')
            if not any(filter_text in str(v).lower() for v in values):
                self.history_tree.detach(item)
            else:
                self.history_tree.reattach(item, '', 'end')

    def create_tooltip(self, widget, text, delay=500):
        toolTip = ToolTip(widget)
        toolTip.delay = delay
        def enter(event):
            ToolTip.showtip(toolTip, text)
        def leave(event):
            ToolTip.hidetip(toolTip)
        widget.bind('<Enter>', enter)
        widget.bind('<Leave>', leave)

    def show_about_window(self):
        if hasattr(self, 'about_window') and self.about_window and self.about_window.winfo_exists():
            self.about_window.lift()
            return

        self.about_window = tk.Toplevel(self.master)
        self.about_window.title("О программе")
        self.about_window.resizable(False, False)
        self.about_window.transient(self.master)

        about_frame = ttk.Frame(self.about_window, padding="20")
        about_frame.pack(fill="both", expand=True)

        app_name_label = ttk.Label(about_frame, text="ScanBox", font=("Segoe UI", 12, "bold"))
        app_name_label.pack(pady=(0, 5))

        version_label = ttk.Label(about_frame, text="Версия 0.9.0.1 BETA Hotfix", font=("Segoe UI", 10))
        version_label.pack(pady=(0, 10))

        try:
            base_path = sys._MEIPASS if hasattr(sys, '_MEIPASS') else os.path.dirname(os.path.abspath(__file__))
            image_path = os.path.join(base_path, "about_image.png")
            about_image = tk.PhotoImage(file=image_path)
            image_label = ttk.Label(about_frame, image=about_image)
            image_label.image = about_image 
            image_label.pack(pady=(0, 10))
        except tk.TclError:
            print("Не удалось загрузить изображение 'about_image.png'")

        description_text = "Программа для быстрого сканирования коробов и товара. Баги ожидаются, фиксы не факт."
        description_label = ttk.Label(about_frame, text=description_text, justify="center", font=("Segoe UI", 9))
        description_label.pack(pady=(0, 15))

        copyright_label = ttk.Label(about_frame, text="© 2025, Holorigg", font=("Segoe UI", 9, "italic"), foreground="#777")
        copyright_label.pack(pady=(0, 0))

        ok_button = ttk.Button(about_frame, text="OK", command=self.about_window.destroy)
        ok_button.pack(pady=(15, 0))
        ok_button.focus_set()

        self.about_window.bind("<Return>", lambda event: self.about_window.destroy())
        self.about_window.bind("<Escape>", lambda event: self.about_window.destroy())

        self.about_window.update_idletasks()
        width = self.about_window.winfo_width()
        height = self.about_window.winfo_height()
        x = self.master.winfo_x() + (self.master.winfo_width() - width) // 2
        y = self.master.winfo_y() + (self.master.winfo_height() - height) // 2
        self.about_window.geometry(f"+{x}+{y}")


class ToolTip:
    def __init__(self, widget):
        self.widget = widget
        self.tipwindow = None
        self.id = None
        self.x = self.y = 0
        self.delay = 500

    def showtip(self, text):
        "Display text in tooltip window"
        self.text = text
        if self.tipwindow or not self.text:
            return
        def show_delayed_tip():
            if not self.tipwindow:
                x, y, cx, cy = self.widget.bbox("insert")
                x = x + self.widget.winfo_rootx() + 5
                y = y + cy + self.widget.winfo_rooty() +5
                self.tipwindow = tw = tk.Toplevel(self.widget)
                tw.wm_overrideredirect(1)
                tw.wm_geometry("+%d+%d" % (x, y))
                label = tk.Label(tw, text=self.text, justify=tk.LEFT,
                      background="#ffffe0", relief=tk.SOLID, borderwidth=1,
                      font=("tahoma", "8", "normal"))
                label.pack(ipadx=1)
        self.id = self.widget.after(self.delay, show_delayed_tip)

    def hidetip(self):
        if self.id:
            self.widget.after_cancel(self.id)
            self.id = None
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.destroy()



root = tk.Tk()
app = BarcodeApp(root)
root.mainloop()
